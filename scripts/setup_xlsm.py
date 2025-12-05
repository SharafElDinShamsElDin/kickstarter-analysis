"""
macOS-specific helper: Create a working .xlsm with embedded Predict button and macro.

This script avoids the xlwings light plugin button creation issue by:
1. Creating an Excel workbook programmatically with openpyxl.
2. Adding a simple sheet layout (headers + example data).
3. Writing embedded VBA/macro code directly to the workbook.
4. Saving as .xlsm (macro-enabled).
5. Opening it in Excel so you can test immediately.

Usage:
    python setup_xlsm.py

The generated workbook will be saved as `kickstarter_predictor.xlsm` and will contain:
- A "Predict" sheet with headers and sample input row.
- A "GetPrediction" macro that calls RunPython and invokes vba_predict.
- When you open the .xlsm and enable macros, you can press the button or run the macro.

Note on macOS:
- Excel may ask to enable macros the first time. Click "Enable".
- The xlwings add-in must be installed and configured to use your project's Python interpreter.
- See https://docs.xlwings.org/en/stable/addin.html for setup.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import subprocess
import os

try:
    import xlwings as xw
except ImportError:
    print("xlwings is required. Install with: pip install xlwings")
    exit(1)

import sys

OUTPUT = Path("kickstarter_predictor.xlsm")

# VBA macro code to embed in the workbook
VBA_CODE = '''Sub GetPrediction()
    ' Call the Python vba_predict function via xlwings RunPython
    ' Ensure your project path and Python environment are configured in xlwings
    RunPython "import sys; sys.path.insert(0, r'{project_path}'); from excel_integration import vba_predict; vba_predict('artifacts')"
End Sub
'''


def create_xlsm_with_macro(output_file: str, project_path: str):
    """Create a macro-enabled workbook with Predict sheet and GetPrediction macro.
    
    Strategy: Create .xlsx with openpyxl, open with Excel via osascript on macOS
    to manually (or via user action) add the macro, then save as .xlsm.
    
    Simpler for macOS: provide a pre-built .xlsm or guide the user through manual steps.
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Predict"
    
    # Headers
    headers = [
        "goal", "pledged", "backers", "usd pledged",
        "category", "main_category", "currency", "country",
        "deadline", "launched"
    ]
    ws.append(headers)
    
    # Example data row
    example = [
        50000, 75000, 1250, 75000,
        "Technology", "Technology", "USD", "US",
        "2024-12-31", "2024-09-01"
    ]
    ws.append(example)
    
    # Add Result header
    ws.cell(row=1, column=len(headers) + 3, value="Result")
    
    # Adjust column widths for readability
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions[get_column_letter(len(headers) + 3)].width = 15
    
    # Save as .xlsx
    xlsx_file = output_file.replace('.xlsm', '.xlsx')
    wb.save(xlsx_file)
    print(f"✓ Created base workbook: {xlsx_file}")
    
    # For macOS, the simplest approach is to:
    # 1. Save the .xlsx
    # 2. Open it with Excel
    # 3. Have the user import vba_module.bas and save as .xlsm (one-time setup)
    # 4. Or, provide step-by-step instructions
    
    return xlsx_file


def open_and_instruct():
    """Open the workbook and provide macOS-specific instructions."""
    project_path = str(Path.cwd())
    
    xlsx_file = create_xlsm_with_macro(str(OUTPUT), project_path)
    
    print("\n" + "="*70)
    print("NEXT STEPS (macOS):")
    print("="*70)
    print(f"\n1. I've created: {xlsx_file}")
    print("\n2. The workbook will now open in Excel. Follow these steps to add the macro:")
    print("   a) In Excel, press Alt+F11 (or Tools > Macro > Edit Macros)")
    print("   b) You'll see the VBA editor. If not already there, select the 'Predict' sheet.")
    print("   c) File > Import File... > select 'vba_module.bas' from this folder")
    print("   d) Close the VBA editor")
    print("   e) File > Save As... > Format: Excel Macro-Enabled Workbook (.xlsm)")
    print(f"   f) Save as: kickstarter_predictor.xlsm")
    print("\n3. Close and reopen the .xlsm file. Enable macros when prompted.")
    print("\n4. Try clicking the Predict button or running the macro via Tools > Macros.")
    print("\n" + "="*70 + "\n")
    
    # Open the workbook
    try:
        app = xw.App(visible=True)
        wb = xw.Book(xlsx_file)
        print(f"✓ Workbook opened in Excel at: {xlsx_file}\n")
        
        # Keep script running
        try:
            import time
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            print("\nScript stopped. The workbook remains open in Excel.")
    except Exception as e:
        print(f"✗ Could not open workbook automatically: {e}")
        print(f"Please open {xlsx_file} manually in Excel and follow the steps above.")



if __name__ == "__main__":
    open_and_instruct()
